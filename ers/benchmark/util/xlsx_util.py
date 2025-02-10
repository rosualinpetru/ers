from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException


class XLSXUtil:
    def __init__(self, filename):
        self.filename = filename
        try:
            self.workbook = load_workbook(self.filename)
        except (FileNotFoundError, InvalidFileException):
            self.workbook = Workbook()
            default_sheet = self.workbook.active
            self.workbook.remove(default_sheet)

    def write_to_page(self, page_name, row):
        if page_name in self.workbook.sheetnames:
            sheet = self.workbook[page_name]
        else:
            sheet = self.workbook.create_sheet(title=page_name)

        sheet.append(row)

        self.workbook.save(self.filename)

    def close(self):
        self.workbook.save(self.filename)
        self.workbook.close()
